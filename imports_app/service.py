import csv
import hashlib
import io
import re
from datetime import date, datetime
from pathlib import Path

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from .models import (
    CorrespondanceColonneImport,
    ErreurImport,
    ImportOfficiel,
    LigneImport,
)
from .repository import (
    CorrespondanceColonneImportRepository,
    ErreurImportRepository,
    ImportOfficielRepository,
    LigneImportRepository,
)


class ChampsAttendusImportService:
    """Catalogue métier des champs attendus selon le type d'import.

    Le frontend affiche ce catalogue pour aider l'utilisateur à faire la
    correspondance entre les colonnes du fichier et les champs FasoIM.
    """

    CHAMPS = {
        ImportOfficiel.TypeSource.BEPC: [
            {"code": "numero_pv", "libelle": "Numéro PV", "obligatoire": True},
            {"code": "nom", "libelle": "Nom", "obligatoire": True},
            {"code": "prenom", "libelle": "Prénom", "obligatoire": True},
            {"code": "date_naissance", "libelle": "Date de naissance", "obligatoire": False},
            {"code": "sexe", "libelle": "Sexe", "obligatoire": False},
            {"code": "etablissement_origine", "libelle": "Établissement d'origine", "obligatoire": False},
        ],
        ImportOfficiel.TypeSource.BAC: [
            {"code": "numero_pv", "libelle": "Numéro PV", "obligatoire": True},
            {"code": "nom", "libelle": "Nom", "obligatoire": True},
            {"code": "prenom", "libelle": "Prénom", "obligatoire": True},
            {"code": "date_naissance", "libelle": "Date de naissance", "obligatoire": False},
            {"code": "sexe", "libelle": "Sexe", "obligatoire": False},
            {"code": "etablissement_origine", "libelle": "Établissement d'origine", "obligatoire": False},
        ],
        ImportOfficiel.TypeSource.CONCOURS: [
            {"code": "numero_recepisse", "libelle": "Numéro de récépissé", "obligatoire": True},
            {"code": "nom", "libelle": "Nom", "obligatoire": True},
            {"code": "prenom", "libelle": "Prénom", "obligatoire": True},
            {"code": "date_naissance", "libelle": "Date de naissance", "obligatoire": False},
            {"code": "sexe", "libelle": "Sexe", "obligatoire": False},
            {"code": "concours", "libelle": "Concours", "obligatoire": False},
        ],
        ImportOfficiel.TypeSource.SELECTIONNES: [
            {"code": "matricule", "libelle": "Matricule", "obligatoire": False},
            {"code": "reference_selection", "libelle": "Référence de sélection", "obligatoire": False},
            {"code": "nom", "libelle": "Nom", "obligatoire": True},
            {"code": "prenom", "libelle": "Prénom", "obligatoire": True},
            {"code": "date_naissance", "libelle": "Date de naissance", "obligatoire": False},
            {"code": "sexe", "libelle": "Sexe", "obligatoire": False},
            {"code": "structure_origine", "libelle": "Structure d'origine", "obligatoire": False},
        ],
        ImportOfficiel.TypeSource.VOLONTAIRES_ACCEPTES: [
            {"code": "code_suivi", "libelle": "Code de suivi", "obligatoire": True},
            {"code": "nom", "libelle": "Nom", "obligatoire": True},
            {"code": "prenom", "libelle": "Prénom", "obligatoire": True},
            {"code": "date_naissance", "libelle": "Date de naissance", "obligatoire": False},
            {"code": "sexe", "libelle": "Sexe", "obligatoire": False},
        ],
    }

    ALIASES = {
        "numero_pv": ["numero pv", "n pv", "n° pv", "pv", "num pv", "numero_pv"],
        "numero_recepisse": ["recepisse", "récépissé", "numero recepisse", "n° recepisse", "numero_recepisse"],
        "code_suivi": ["code suivi", "code_suivi", "code volontaire", "code inscription"],
        "matricule": ["matricule", "matricule candidat"],
        "reference_selection": ["reference", "référence", "reference selection", "référence sélection"],
        "nom": ["nom", "nom candidat", "nom de famille"],
        "prenom": ["prenom", "prénom", "prenoms", "prénoms", "prenom candidat", "prénom candidat"],
        "date_naissance": ["date naissance", "date de naissance", "né le", "ne le", "date nais"],
        "sexe": ["sexe", "genre", "civilite", "civilité"],
        "etablissement_origine": ["etablissement", "établissement", "ecole", "école", "lycee", "lycée"],
        "structure_origine": ["structure", "structure origine", "service", "direction"],
        "concours": ["concours", "type concours", "emploi", "poste"],
    }

    @classmethod
    def lister(cls, type_source):
        cls.valider_type_source(type_source)
        return list(cls.CHAMPS[type_source])

    @classmethod
    def codes(cls, type_source):
        return [champ["code"] for champ in cls.lister(type_source)]

    @classmethod
    def obligatoires(cls, type_source):
        return [champ["code"] for champ in cls.lister(type_source) if champ.get("obligatoire")]

    @classmethod
    def dictionnaire(cls, type_source):
        return {champ["code"]: champ for champ in cls.lister(type_source)}

    @classmethod
    def valider_type_source(cls, type_source):
        if type_source not in cls.CHAMPS:
            raise ValidationError({"type_source": "Type de source d'import non pris en charge."})

    @staticmethod
    def normaliser_libelle(valeur):
        valeur = "" if valeur is None else str(valeur)
        valeur = valeur.strip().lower()
        valeur = re.sub(r"[\n\r\t]+", " ", valeur)
        valeur = re.sub(r"[_\-]+", " ", valeur)
        valeur = re.sub(r"\s+", " ", valeur)
        return valeur

    @classmethod
    def champ_probable_pour_colonne(cls, colonne, type_source):
        colonne_normale = cls.normaliser_libelle(colonne)
        if not colonne_normale:
            return None

        for code in cls.codes(type_source):
            alias = cls.ALIASES.get(code, []) + [code]
            for element in alias:
                alias_normalise = cls.normaliser_libelle(element)
                if colonne_normale == alias_normalise or alias_normalise in colonne_normale:
                    return code
        return None

    @classmethod
    def score_ligne_entete(cls, cellules, type_source):
        cellules_utiles = [str(cellule).strip() for cellule in cellules if str(cellule or "").strip()]
        if len(cellules_utiles) < 2:
            return 0

        champs_detectes = set()
        score = len(cellules_utiles)

        for cellule in cellules_utiles:
            cellule_normale = cls.normaliser_libelle(cellule)
            if len(cellule_normale) > 70 and not any(mot in cellule_normale for mot in ["nom", "pv", "sexe", "date"]):
                score -= 2
            champ = cls.champ_probable_pour_colonne(cellule, type_source)
            if champ:
                champs_detectes.add(champ)
                score += 5

        obligatoires_detectes = champs_detectes.intersection(set(cls.obligatoires(type_source)))
        score += len(obligatoires_detectes) * 4
        return max(score, 0)


class LectureFichierImportService:
    """Lecture légère des fichiers d'import.

    On tente d'abord de lire le début du fichier comme tableau. Si ce n'est pas
    convaincant, on scanne les premières lignes pour trouver l'entête probable.
    Les images, logos et textes fixes ne sont pas stockés.
    """

    NB_LIGNES_SCAN = 50
    NB_LIGNES_APERCU = 0
    DELIMITEURS_CSV = [";", ",", "|", "\t"]
    EXTENSIONS_EXCEL = {".xlsx"}
    EXTENSIONS_CSV = {".csv"}

    @classmethod
    def detecter_type_fichier(cls, nom_fichier):
        extension = Path(nom_fichier or "").suffix.lower()
        if extension in cls.EXTENSIONS_EXCEL:
            return ImportOfficiel.TypeFichier.EXCEL
        if extension in cls.EXTENSIONS_CSV:
            return ImportOfficiel.TypeFichier.CSV
        raise ValidationError({"fichier": "Format de fichier non pris en charge. Utilisez .xlsx ou .csv."})

    @classmethod
    def analyser_structure(cls, import_officiel):
        if import_officiel.type_fichier == ImportOfficiel.TypeFichier.EXCEL:
            return cls._analyser_excel(import_officiel)
        if import_officiel.type_fichier == ImportOfficiel.TypeFichier.CSV:
            return cls._analyser_csv(import_officiel)
        raise ValidationError({"type_fichier": "Type de fichier non pris en charge."})

    @classmethod
    def _analyser_excel(cls, import_officiel):
        try:
            from openpyxl import load_workbook
        except ImportError as erreur:
            raise ValidationError({"fichier": "La dépendance openpyxl est nécessaire pour lire les fichiers .xlsx."}) from erreur

        meilleurs = []
        with import_officiel.fichier.open("rb") as fichier:
            classeur = load_workbook(filename=fichier, read_only=True, data_only=True)
            for feuille in classeur.worksheets:
                if getattr(feuille, "sheet_state", "visible") != "visible":
                    continue
                lignes = list(
                    feuille.iter_rows(
                        min_row=1,
                        max_row=cls.NB_LIGNES_SCAN,
                        values_only=True,
                    )
                )
                resultat = cls._trouver_entete_probable(lignes, import_officiel.type_source)
                if resultat:
                    resultat["feuille"] = feuille.title
                    meilleurs.append(resultat)
            classeur.close()

        if not meilleurs:
            raise ValidationError({"fichier": "Aucun tableau exploitable n'a été détecté dans le fichier Excel."})

        meilleur = sorted(meilleurs, key=lambda item: item["score_entete"], reverse=True)[0]
        return {
            "colonnes_detectees": meilleur["colonnes"],
            "apercu_lignes": [],
            "parametres_lecture": {
                "type_fichier": "xlsx",
                "feuille": meilleur["feuille"],
                "ligne_entete": meilleur["ligne_entete"],
                "premiere_ligne_donnees": meilleur["premiere_ligne_donnees"],
                "mode_detection": meilleur["mode_detection"],
                "score_entete": meilleur["score_entete"],
                "nb_lignes_analysees": cls.NB_LIGNES_SCAN,
            },
        }

    @classmethod
    def _analyser_csv(cls, import_officiel):
        texte = cls._lire_csv_texte(import_officiel)
        separateur = cls._detecter_separateur_csv(texte)
        lecteur = csv.reader(io.StringIO(texte), delimiter=separateur)
        lignes = []
        for index, ligne in enumerate(lecteur, start=1):
            lignes.append(ligne)
            if index >= cls.NB_LIGNES_SCAN:
                break

        resultat = cls._trouver_entete_probable(lignes, import_officiel.type_source)
        if not resultat:
            raise ValidationError({"fichier": "Aucun tableau exploitable n'a été détecté dans le fichier CSV."})

        return {
            "colonnes_detectees": resultat["colonnes"],
            "apercu_lignes": [],
            "parametres_lecture": {
                "type_fichier": "csv",
                "separateur_csv": separateur,
                "ligne_entete": resultat["ligne_entete"],
                "premiere_ligne_donnees": resultat["premiere_ligne_donnees"],
                "mode_detection": resultat["mode_detection"],
                "score_entete": resultat["score_entete"],
                "nb_lignes_analysees": cls.NB_LIGNES_SCAN,
            },
        }

    @classmethod
    def _trouver_entete_probable(cls, lignes, type_source):
        lignes_utiles = []
        for numero, ligne in enumerate(lignes, start=1):
            cellules = cls._normaliser_ligne(ligne)
            if any(cellules):
                lignes_utiles.append((numero, cellules))

        if not lignes_utiles:
            return None

        seuil = cls._seuil_entete(type_source)
        numero_debut, ligne_debut = lignes_utiles[0]
        score_debut = ChampsAttendusImportService.score_ligne_entete(ligne_debut, type_source)

        if score_debut >= seuil:
            return cls._resultat_entete(numero_debut, ligne_debut, score_debut, "debut_fichier")

        meilleur = None
        for numero, ligne in lignes_utiles:
            score = ChampsAttendusImportService.score_ligne_entete(ligne, type_source)
            if meilleur is None or score > meilleur["score_entete"]:
                meilleur = cls._resultat_entete(numero, ligne, score, "scan")

        if meilleur and meilleur["score_entete"] >= seuil:
            return meilleur
        return None

    @staticmethod
    def _normaliser_ligne(ligne):
        return ["" if cellule is None else str(cellule).strip() for cellule in list(ligne or [])]

    @staticmethod
    def _resultat_entete(numero_ligne, ligne, score, mode_detection):
        colonnes = []
        deja_vues = {}
        for cellule in ligne:
            colonne = str(cellule or "").strip()
            if not colonne:
                continue
            compteur = deja_vues.get(colonne, 0)
            deja_vues[colonne] = compteur + 1
            if compteur:
                colonne = f"{colonne} ({compteur + 1})"
            colonnes.append(colonne)

        return {
            "ligne_entete": numero_ligne,
            "premiere_ligne_donnees": numero_ligne + 1,
            "colonnes": colonnes,
            "score_entete": score,
            "mode_detection": mode_detection,
        }

    @staticmethod
    def _seuil_entete(type_source):
        obligatoires = ChampsAttendusImportService.obligatoires(type_source)
        return 10 + (len(obligatoires) * 2)

    @staticmethod
    def _lire_csv_texte(import_officiel):
        with import_officiel.fichier.open("rb") as fichier:
            contenu = fichier.read()
        for encodage in ["utf-8-sig", "utf-8", "latin-1"]:
            try:
                return contenu.decode(encodage)
            except UnicodeDecodeError:
                continue
        raise ValidationError({"fichier": "Impossible de lire l'encodage du fichier CSV."})

    @classmethod
    def _detecter_separateur_csv(cls, texte):
        echantillon = texte[:4096]
        try:
            dialecte = csv.Sniffer().sniff(echantillon, delimiters="".join(cls.DELIMITEURS_CSV))
            return dialecte.delimiter
        except csv.Error:
            scores = {delimiteur: echantillon.count(delimiteur) for delimiteur in cls.DELIMITEURS_CSV}
            return max(scores, key=scores.get) if any(scores.values()) else ";"

    @classmethod
    def iterer_lignes_normalisees(cls, import_officiel, mapping):
        if import_officiel.type_fichier == ImportOfficiel.TypeFichier.EXCEL:
            yield from cls._iterer_lignes_excel(import_officiel, mapping)
            return
        if import_officiel.type_fichier == ImportOfficiel.TypeFichier.CSV:
            yield from cls._iterer_lignes_csv(import_officiel, mapping)
            return
        raise ValidationError({"type_fichier": "Type de fichier non pris en charge."})

    @classmethod
    def _iterer_lignes_excel(cls, import_officiel, mapping):
        try:
            from openpyxl import load_workbook
        except ImportError as erreur:
            raise ValidationError({"fichier": "La dépendance openpyxl est nécessaire pour lire les fichiers .xlsx."}) from erreur

        parametres = import_officiel.parametres_lecture or {}
        feuille_nom = parametres.get("feuille")
        ligne_entete = int(parametres.get("ligne_entete") or 1)
        premiere_ligne = int(parametres.get("premiere_ligne_donnees") or ligne_entete + 1)

        with import_officiel.fichier.open("rb") as fichier:
            classeur = load_workbook(filename=fichier, read_only=True, data_only=True)
            feuille = classeur[feuille_nom] if feuille_nom in classeur.sheetnames else classeur.worksheets[0]
            entete = next(
                feuille.iter_rows(min_row=ligne_entete, max_row=ligne_entete, values_only=True),
                [],
            )
            index_colonnes = cls._index_colonnes(entete)
            for numero_ligne, ligne in enumerate(
                feuille.iter_rows(min_row=premiere_ligne, values_only=True),
                start=premiere_ligne,
            ):
                resultat = cls._normaliser_ligne_donnees(numero_ligne, ligne, mapping, index_colonnes)
                if resultat:
                    yield resultat
            classeur.close()

    @classmethod
    def _iterer_lignes_csv(cls, import_officiel, mapping):
        texte = cls._lire_csv_texte(import_officiel)
        parametres = import_officiel.parametres_lecture or {}
        separateur = parametres.get("separateur_csv") or cls._detecter_separateur_csv(texte)
        ligne_entete = int(parametres.get("ligne_entete") or 1)
        premiere_ligne = int(parametres.get("premiere_ligne_donnees") or ligne_entete + 1)

        lecteur = list(csv.reader(io.StringIO(texte), delimiter=separateur))
        entete = lecteur[ligne_entete - 1] if len(lecteur) >= ligne_entete else []
        index_colonnes = cls._index_colonnes(entete)

        for numero_ligne, ligne in enumerate(lecteur[premiere_ligne - 1 :], start=premiere_ligne):
            resultat = cls._normaliser_ligne_donnees(numero_ligne, ligne, mapping, index_colonnes)
            if resultat:
                yield resultat

    @staticmethod
    def _index_colonnes(entete):
        index = {}
        for position, colonne in enumerate(entete):
            libelle = "" if colonne is None else str(colonne).strip()
            if libelle and libelle not in index:
                index[libelle] = position
        return index

    @classmethod
    def _normaliser_ligne_donnees(cls, numero_ligne, ligne, mapping, index_colonnes):
        cellules = list(ligne or [])
        if not any(str(cellule or "").strip() for cellule in cellules):
            return None

        donnees_brutes = {}
        donnees_normalisees = {}
        for champ_cible, colonne_source in mapping.items():
            position = index_colonnes.get(colonne_source)
            valeur = cellules[position] if position is not None and position < len(cellules) else ""
            donnees_brutes[colonne_source] = cls._serialiser_valeur(valeur)
            donnees_normalisees[champ_cible] = cls._normaliser_valeur(champ_cible, valeur)

        return numero_ligne, donnees_brutes, donnees_normalisees

    @staticmethod
    def _serialiser_valeur(valeur):
        if isinstance(valeur, (datetime, date)):
            return valeur.isoformat()
        return "" if valeur is None else str(valeur).strip()

    @classmethod
    def _normaliser_valeur(cls, champ_cible, valeur):
        valeur = cls._serialiser_valeur(valeur)
        if champ_cible in {"nom", "prenom"}:
            return re.sub(r"\s+", " ", valeur).strip().upper()
        if champ_cible == "sexe":
            valeur_normale = valeur.strip().lower()
            if valeur_normale in {"m", "masculin", "homme", "garcon", "garçon"}:
                return "M"
            if valeur_normale in {"f", "feminin", "féminin", "femme", "fille"}:
                return "F"
            return valeur.strip().upper()
        return valeur.strip()


class ImportOfficielService:
    """Service métier des dossiers d'import officiel."""

    @staticmethod
    def calculer_hash_et_taille(fichier):
        position = None
        if hasattr(fichier, "tell") and hasattr(fichier, "seek"):
            try:
                position = fichier.tell()
            except (OSError, ValueError):
                position = None

        hash_sha256 = hashlib.sha256()
        taille = 0
        chunks = fichier.chunks() if hasattr(fichier, "chunks") else iter(lambda: fichier.read(8192), b"")
        for chunk in chunks:
            if isinstance(chunk, str):
                chunk = chunk.encode("utf-8")
            taille += len(chunk)
            hash_sha256.update(chunk)

        if position is not None:
            try:
                fichier.seek(position)
            except (OSError, ValueError):
                pass
        elif hasattr(fichier, "seek"):
            try:
                fichier.seek(0)
            except (OSError, ValueError):
                pass

        return hash_sha256.hexdigest(), taille

    @staticmethod
    @transaction.atomic
    def creer_import_officiel(*, session, type_source, fichier, importe_par=None, commentaire="", lancer_async=True):
        ChampsAttendusImportService.valider_type_source(type_source)
        nom_original = Path(getattr(fichier, "name", "") or "import").name
        type_fichier = LectureFichierImportService.detecter_type_fichier(nom_original)
        hash_fichier, taille_fichier = ImportOfficielService.calculer_hash_et_taille(fichier)

        import_officiel = ImportOfficiel.objects.create(
            session=session,
            type_source=type_source,
            type_fichier=type_fichier,
            fichier=fichier,
            nom_fichier_original=nom_original,
            taille_fichier=taille_fichier,
            hash_fichier=hash_fichier,
            commentaire=commentaire or "",
            importe_par=importe_par,
            statut=ImportOfficiel.Statut.RECU,
        )

        if lancer_async:
            ImportOfficielService.planifier_lecture_colonnes(import_officiel)
        return import_officiel

    @staticmethod
    def planifier_lecture_colonnes(import_officiel):
        try:
            from .tasks import lire_colonnes_import_task
        except ModuleNotFoundError:
            return False
        lire_colonnes_import_task.delay(import_officiel.id)
        return True

    @staticmethod
    def planifier_validation_lignes(import_officiel):
        try:
            from .tasks import valider_lignes_import_task
        except ModuleNotFoundError:
            return False
        valider_lignes_import_task.delay(import_officiel.id)
        return True

    @staticmethod
    def analyser_colonnes(import_id):
        import_officiel = ImportOfficielRepository.get_by_id(import_id)
        if not import_officiel:
            raise ValidationError({"import": "Import officiel introuvable."})

        ImportOfficielRepository.mettre_a_jour_statut(
            import_officiel,
            ImportOfficiel.Statut.LECTURE_COLONNES_EN_COURS,
            message="",
        )

        try:
            analyse = LectureFichierImportService.analyser_structure(import_officiel)
        except ValidationError as erreur:
            ImportOfficielRepository.mettre_a_jour_statut(
                import_officiel,
                ImportOfficiel.Statut.ECHEC,
                message=str(erreur),
            )
            raise

        with transaction.atomic():
            import_verrouille = ImportOfficielRepository.get_by_id_pour_update(import_id)
            if not import_verrouille:
                raise ValidationError({"import": "Import officiel introuvable."})
            import_verrouille.colonnes_detectees = analyse["colonnes_detectees"]
            import_verrouille.apercu_lignes = analyse.get("apercu_lignes", [])
            import_verrouille.parametres_lecture = analyse.get("parametres_lecture", {})
            import_verrouille.statut = ImportOfficiel.Statut.CORRESPONDANCE_REQUISE
            import_verrouille.message_erreur = ""
            import_verrouille.date_lecture_colonnes = timezone.now()
            import_verrouille.updated_at = timezone.now()
            import_verrouille.save(
                update_fields=[
                    "colonnes_detectees",
                    "apercu_lignes",
                    "parametres_lecture",
                    "statut",
                    "message_erreur",
                    "date_lecture_colonnes",
                    "updated_at",
                ]
            )
            return import_verrouille

    @staticmethod
    @transaction.atomic
    def annuler(import_id, acteur=None, message=""):
        import_officiel = ImportOfficielRepository.get_by_id_pour_update(import_id)
        if not import_officiel:
            raise ValidationError({"import": "Import officiel introuvable."})
        import_officiel.statut = ImportOfficiel.Statut.ANNULE
        import_officiel.message_erreur = message or "Import annulé."
        import_officiel.date_fin_traitement = timezone.now()
        import_officiel.updated_at = timezone.now()
        import_officiel.save(update_fields=["statut", "message_erreur", "date_fin_traitement", "updated_at"])
        return import_officiel


class CorrespondanceColonneImportService:
    """Service métier de la correspondance des colonnes."""

    @staticmethod
    @transaction.atomic
    def valider_correspondance(*, import_id, correspondances, confirme_par=None, parametres_lecture=None, lancer_async=True):
        import_officiel = ImportOfficielRepository.get_by_id_pour_update(import_id)
        if not import_officiel:
            raise ValidationError({"import": "Import officiel introuvable."})
        if not import_officiel.peut_recevoir_correspondance:
            raise ValidationError({"statut": "Cet import ne peut pas recevoir de correspondance maintenant."})

        colonnes_detectees = set(import_officiel.colonnes_detectees or [])
        champs = ChampsAttendusImportService.dictionnaire(import_officiel.type_source)
        champs_valides = set(champs.keys())
        cibles_vues = set()
        colonnes_vues = set()
        erreurs = []
        objets = []

        for ordre, correspondance in enumerate(correspondances or [], start=1):
            champ_cible = str(correspondance.get("champ_cible", "")).strip()
            colonne_source = str(correspondance.get("colonne_source", "")).strip()
            if champ_cible not in champs_valides:
                erreurs.append(f"Champ cible invalide : {champ_cible}")
            if colonne_source not in colonnes_detectees:
                erreurs.append(f"Colonne source inconnue : {colonne_source}")
            if champ_cible in cibles_vues:
                erreurs.append(f"Champ cible utilisé plusieurs fois : {champ_cible}")
            if colonne_source in colonnes_vues:
                erreurs.append(f"Colonne source utilisée plusieurs fois : {colonne_source}")
            if erreurs:
                continue

            cibles_vues.add(champ_cible)
            colonnes_vues.add(colonne_source)
            objets.append(
                CorrespondanceColonneImport(
                    import_officiel=import_officiel,
                    champ_cible=champ_cible,
                    libelle_champ_cible=champs[champ_cible]["libelle"],
                    colonne_source=colonne_source,
                    obligatoire=bool(champs[champ_cible].get("obligatoire")),
                    confirmee=True,
                    ordre=ordre,
                )
            )

        champs_obligatoires = set(ChampsAttendusImportService.obligatoires(import_officiel.type_source))
        champs_absents = champs_obligatoires.difference(cibles_vues)
        if champs_absents:
            erreurs.append("Champs obligatoires non associés : " + ", ".join(sorted(champs_absents)))

        if import_officiel.type_source == ImportOfficiel.TypeSource.SELECTIONNES:
            if not ({"matricule", "reference_selection"}.intersection(cibles_vues)):
                erreurs.append("Pour les sélectionnés, matricule ou référence de sélection doit être associé.")

        if erreurs:
            raise ValidationError({"correspondances": erreurs})

        CorrespondanceColonneImportRepository.supprimer_logiquement_par_import(import_officiel)
        CorrespondanceColonneImportRepository.creer_en_masse(objets)

        nouveaux_parametres = dict(import_officiel.parametres_lecture or {})
        if parametres_lecture:
            nouveaux_parametres.update(parametres_lecture)
        if nouveaux_parametres.get("ligne_entete") and not nouveaux_parametres.get("premiere_ligne_donnees"):
            nouveaux_parametres["premiere_ligne_donnees"] = int(nouveaux_parametres["ligne_entete"]) + 1

        import_officiel.parametres_lecture = nouveaux_parametres
        import_officiel.correspondance_confirmee_par = confirme_par
        import_officiel.statut = ImportOfficiel.Statut.CORRESPONDANCE_VALIDEE
        import_officiel.date_correspondance = timezone.now()
        import_officiel.updated_at = timezone.now()
        import_officiel.save(
            update_fields=[
                "parametres_lecture",
                "correspondance_confirmee_par",
                "statut",
                "date_correspondance",
                "updated_at",
            ]
        )

        if lancer_async:
            ImportOfficielService.planifier_validation_lignes(import_officiel)
        return import_officiel


class ValidationImportService:
    """Validation des lignes après correspondance des colonnes."""

    BATCH_SIZE = 1000

    @classmethod
    def valider_lignes(cls, import_id):
        import_officiel = ImportOfficielRepository.get_by_id(import_id)
        if not import_officiel:
            raise ValidationError({"import": "Import officiel introuvable."})
        if import_officiel.statut not in {
            ImportOfficiel.Statut.CORRESPONDANCE_VALIDEE,
            ImportOfficiel.Statut.VALIDATION_EN_COURS,
            ImportOfficiel.Statut.VALIDE_AVEC_ERREURS,
        }:
            raise ValidationError({"statut": "La correspondance doit être validée avant la validation des lignes."})

        ImportOfficielRepository.mettre_a_jour_statut(import_officiel, ImportOfficiel.Statut.VALIDATION_EN_COURS, message="")
        mapping = CorrespondanceColonneImportRepository.mapping_par_import(import_officiel, confirmees_seulement=True)
        if not mapping:
            raise ValidationError({"correspondances": "Aucune correspondance confirmée."})

        LigneImportRepository.supprimer_logiquement_par_import(import_officiel)
        ErreurImportRepository.supprimer_logiquement_par_import(import_officiel)

        lignes_batch = []
        erreurs_batch = []
        lignes_creees_total = 0
        erreurs_total = 0

        for numero_ligne, donnees_brutes, donnees_normalisees in LectureFichierImportService.iterer_lignes_normalisees(import_officiel, mapping):
            erreurs_ligne = cls._valider_donnees(import_officiel.type_source, donnees_normalisees)
            statut = LigneImport.Statut.ERREUR if erreurs_ligne else LigneImport.Statut.VALIDE
            ligne = LigneImport(
                import_officiel=import_officiel,
                numero_ligne=numero_ligne,
                donnees_brutes=donnees_brutes,
                donnees_normalisees=donnees_normalisees,
                statut=statut,
                message_statut="Erreur détectée." if erreurs_ligne else "Ligne valide.",
                hash_ligne=cls._hash_ligne(donnees_normalisees),
            )
            lignes_batch.append((ligne, erreurs_ligne))
            if len(lignes_batch) >= cls.BATCH_SIZE:
                creees, erreurs = cls._enregistrer_lignes_batch(import_officiel, lignes_batch, mapping)
                lignes_creees_total += creees
                erreurs_total += erreurs
                lignes_batch = []

        if lignes_batch:
            creees, erreurs = cls._enregistrer_lignes_batch(import_officiel, lignes_batch, mapping)
            lignes_creees_total += creees
            erreurs_total += erreurs

        import_officiel = ImportOfficielRepository.mettre_a_jour_statistiques(import_officiel)
        statut_final = ImportOfficiel.Statut.VALIDE_AVEC_ERREURS if erreurs_total else ImportOfficiel.Statut.VALIDE
        ImportOfficielRepository.mettre_a_jour_statut(import_officiel, statut_final, message="")
        return import_officiel

    @classmethod
    def _enregistrer_lignes_batch(cls, import_officiel, lignes_batch, mapping):
        lignes = [element[0] for element in lignes_batch]
        lignes_creees = LigneImportRepository.creer_en_masse(lignes, batch_size=cls.BATCH_SIZE)
        erreurs = []
        for ligne, erreurs_ligne in zip(lignes_creees, [element[1] for element in lignes_batch]):
            for erreur in erreurs_ligne:
                colonne_source = mapping.get(erreur["champ_cible"], "")
                erreurs.append(
                    ErreurImport(
                        import_officiel=import_officiel,
                        ligne_import=ligne,
                        champ_cible=erreur["champ_cible"],
                        colonne_source=colonne_source,
                        type_erreur=erreur["type_erreur"],
                        gravite=erreur["gravite"],
                        message=erreur["message"],
                        valeur_recue=erreur.get("valeur_recue", ""),
                        code_erreur=erreur.get("code_erreur", ""),
                    )
                )
        if erreurs:
            ErreurImportRepository.creer_en_masse(erreurs, batch_size=cls.BATCH_SIZE)
        return len(lignes_creees), len(erreurs)

    @classmethod
    def _valider_donnees(cls, type_source, donnees):
        erreurs = []
        for champ in ChampsAttendusImportService.obligatoires(type_source):
            if not str(donnees.get(champ, "")).strip():
                erreurs.append(cls._erreur(champ, ErreurImport.TypeErreur.CHAMP_OBLIGATOIRE, "Champ obligatoire manquant."))

        if type_source == ImportOfficiel.TypeSource.SELECTIONNES:
            if not str(donnees.get("matricule", "")).strip() and not str(donnees.get("reference_selection", "")).strip():
                erreurs.append(
                    cls._erreur(
                        "matricule",
                        ErreurImport.TypeErreur.CHAMP_OBLIGATOIRE,
                        "Matricule ou référence de sélection obligatoire.",
                    )
                )

        sexe = str(donnees.get("sexe", "")).strip()
        if sexe and sexe not in {"M", "F"}:
            erreurs.append(
                cls._erreur(
                    "sexe",
                    ErreurImport.TypeErreur.VALEUR_INVALIDE,
                    "Le sexe doit être M ou F.",
                    valeur_recue=sexe,
                )
            )
        return erreurs

    @staticmethod
    def _erreur(champ_cible, type_erreur, message, valeur_recue=""):
        return {
            "champ_cible": champ_cible,
            "type_erreur": type_erreur,
            "gravite": ErreurImport.Gravite.BLOQUANTE,
            "message": message,
            "valeur_recue": valeur_recue,
            "code_erreur": f"{champ_cible}.{type_erreur}".lower(),
        }

    @staticmethod
    def _hash_ligne(donnees_normalisees):
        contenu = "|".join(f"{cle}={donnees_normalisees.get(cle, '')}" for cle in sorted(donnees_normalisees))
        return hashlib.sha256(contenu.encode("utf-8")).hexdigest()
